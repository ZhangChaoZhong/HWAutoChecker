import docx2txt
import os.path
import glob
import re
import hashlib
import LabReport
from gensim import corpora, models, similarities
import pandas as pd
import numpy as np
from scipy import stats
from collections import Counter
import matplotlib.pyplot as plt
import openpyxl as pyxl

datafile = '/Users/wliu/PycharmProjects/HWAutoChecker/data/云平台管理技术/实验1'
textTags = ['实验步骤', '实验总结']
grade = {'nImages': {'A': 95, 'B': 85, 'C': 75, 'D': 70, 'E': 60},
         'nKeywords': {
             '实验步骤': {'A': 95, 'B': 85, 'C': 75, 'D': 70, 'E': 60},
             '实验总结': {'A': 95, 'B': 85, 'C': 75, 'D': 70, 'E': 60},
         }}
weight = {'nImages': 0.7,
          'nKeywords': {
              '实验步骤': 0.1,
              '实验总结': 0.2,
          }}
remarks = {'nImages': {
    'A': '实验结果记录非常详细，',
    'B': '实验结果记录比较详细，',
    'C': '实验结果记录基本完整，',
    'D': '实验结果记录部分完整，',
    'E': '实验结果记录欠完整，'
},
    'nKeywords': {
        '实验步骤': {
            'A': '实验步骤正确完整，',
            'B': '实验步骤比较完整，',
            'C': '实验步骤基本完整，',
            'D': '实验步骤欠完整，',
            'E': '实验步骤缺失比较多，'
        },
        '实验总结': {
            'A': '实验总结详尽且到位。',
            'B': '实验总结比较到位。',
            'C': '实验总结基本到位。',
            'D': '实验总结不到位。',
            'E': '实验总结缺。'
        },
    }}


def calculateTextSimilarity(reports, tag, threshold=0.8):
    corpus = []
    for report in reports.values():
        corpus.append(report.keywords[tag])

    # 以下使用doc2bow制作语料库 使用TF-IDF模型对语料库建模
    dictionary = corpora.Dictionary(corpus)
    bow_corpus = [dictionary.doc2bow(doc) for doc in corpus]
    tfidf = models.TfidfModel(bow_corpus)
    corpus_tfidf = tfidf[bow_corpus]
    similar_matrix = similarities.MatrixSimilarity(corpus_tfidf)

    # compare
    for id, report in reports.items():
        corpus = dictionary.doc2bow(report.keywords[tag])
        test_tfidf = tfidf[corpus]
        result = similar_matrix[test_tfidf]
        for value, (_id, _report) in zip(result, reports.items()):
            if value > threshold and id != _id:
                report.appendSimilarityText(tag, {'studentID': _report.studentID, 'studentName': _report.studentName,
                                                  'similarity': value})
                # print(report.studentID, report.studentName, "-- 相似度 {:.2f}".format(value), _report.studentID, _report.studentName)


def calculateImageSimilarity(reports, threshold=0.2):
    for id, report in reports.items():
        imgfileMd5 = report.imgfileMd5
        imgSimilarity = [len(imgfileMd5.intersection(_report.imgfileMd5)) / len(report.imgfileMd5) if len(
            report.imgfileMd5) > 0 and id != _id else 0 for _id, _report in reports.items()]
        # get similarity info
        for value, (_id, _report) in zip(imgSimilarity, reports.items()):
            if value > threshold and id != _id:
                report.appendSimilarityImage({'studentID': _report.studentID, 'studentName': _report.studentName,
                                              'similarity': value})


def drawHistgram(data):
    data_pd = pd.Series(data)
    l_unique_data = list(data_pd.value_counts().index)  # 该数据中的唯一值
    l_num = list(data_pd.value_counts())  # 唯一值出现的次数
    plt.bar(l_unique_data, l_num, width=0.1)
    plt.show()


def getSectionPoint(dataArray):
    x = np.array(dataArray)
    mean, std = x.mean(), x.std(ddof=1)
    a90 = stats.norm.interval(0.8, loc=mean, scale=std)
    a80 = stats.norm.interval(0.5, loc=mean, scale=std)
    return (0, a90[0], a80[0], a80[1], a90[1], x.max())


def statisticData2Score(statisticData):
    sections = getSectionPoint(statisticData)
    result = pd.cut(statisticData, sections, labels=['E', 'D', 'C', 'B', 'A'])
    return result


def traversal(path):
    files = [f for f in glob.glob(path + "**/*.docx", recursive=True)]
    reports = {}
    reportsFailed = {}
    # step1: get raw text from files
    for f in files:
        report = LabReport.LabReport(f, textTags)
        if report.status == 'ok':
            reports[report.studentID] = report
        else:
            reportsFailed[report.studentID] = report

    # step2: calculate similarity
    calculateImageSimilarity(reports)
    for tag in textTags:
        calculateTextSimilarity(reports, tag)

    print("total number of files:", len(files))
    print("total number of correctly parsed records:", len(reports))
    print("total number of failed to parse records:", len(reportsFailed))
    for report in reportsFailed.values():
        print(report.filename, report.status)

    # step3: set the scores for each sections
    result = dict()
    statisticText = dict()
    statisticImages = [report.statistics['nImages'] for report in reports.values()]
    result["images"] = statisticData2Score(statisticImages)
    result["text"] = []
    textResult = dict()
    for tag in textTags:
        statisticText[tag] = [report.statistics['nKeywords'][tag] for report in reports.values()]
        textResult[tag] = statisticData2Score(statisticText[tag])
        # print(textResult[tag])

    for i in range(len(reports)):
        resultEntity = dict()
        for tag in textTags:
            resultEntity[tag] = textResult[tag][i]
        result['text'].append(resultEntity)
        # print(resultEntity)

    # step4: save scores to each report
    for report, img, text, in zip(reports.values(), result["images"], result["text"]):
        # print(report.studentID, report.studentName, report.statistics["nImages"],
        #       report.statistics['nKeywords']["实验步骤"], report.statistics['nKeywords']["实验总结"], \
        #       img, text)
        report.scores['nImages'] = img
        report.scores['nKeywords'] = text
        report.evaluate(grade, weight, remarks, textTags)
    # drawHistgram(statisticText[textTags[1]])
    return reports


def saveReportInfo(filename, sheetname, reports):
    wb = pyxl.load_workbook(filename)
    if sheetname in wb.sheetnames:
        ws = wb[sheetname]
    else:
        ws = wb.copy_worksheet(wb[wb.sheetnames[0]])
    ws.title = sheetname
    headIndex = ord('D')
    for i in range(2):
        colIndex = chr(headIndex) + '1'
        ws[colIndex] = '图片数量'
        headIndex += 1
        for tag in textTags:
            colIndex = chr(headIndex) + '1'
            ws[colIndex] = tag
            headIndex += 1
    colIndex = chr(headIndex) + '1'
    ws[colIndex] = '成绩'
    headIndex += 1
    colIndex = chr(headIndex) + '1'
    ws[colIndex] = '评语'
    headIndex += 1

    for row in ws.iter_rows(min_row=2):
        studentID = row[0].value
        colIndex = 3
        if studentID in reports:
            report = reports[row[0].value]
            reports[row[0].value].print()
            row[colIndex].value = report.statistics["nImages"]
            colIndex += 1
            for tag in textTags:
                row[colIndex].value = report.statistics["nKeywords"][tag]
                colIndex += 1

            row[colIndex].value = report.scores["nImages"]
            colIndex += 1
            for tag in textTags:
                row[colIndex].value = report.scores["nKeywords"][tag]
                colIndex += 1
            row[colIndex].value = report.scores["final"]
            colIndex += 1
            row[colIndex].value = report.scores["remark"]
            colIndex += 1

        for i in row:
            print(i.value, end="\t")

        print(report.scores['final'])

    wb.save(filename)
    wb.close()


def main():
    reports = traversal(datafile)
    xlsFileName = "/Users/wliu/PycharmProjects/HWAutoChecker/data/云平台管理技术/成绩登记表.xlsx"
    saveReportInfo(xlsFileName, '实验1', reports)


if __name__ == "__main__":
    main()
